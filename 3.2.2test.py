import math
from _datetime import datetime
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader
import pdfkit
from openpyxl.reader.excel import load_workbook
import pandas as pd
import os
import multiprocessing

"""Словарь для перевода з/п в рубли"""
currency_to_rub = {
    "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
    "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}


class DataSet:
    """
    Класс для хранения списка вакансий.

    Attributes:
        file_name (str): Название файла
        vacancies_objects (list): Список вакансий
    """

    def __init__(self, file_name):
        """
        Конструктор для инициализация объекта DataSet, который создает поле для хранения списка вакансий

        Args:
             file_name (str): Название файла
        """
        self.file_name = file_name
        self.salary_by_year = dict()
        self.vacancies_count_by_year = dict()
        self.salary_by_profession_name = dict()
        self.vacancies_count_by_profession_name = dict()
        self.salary_by_city = dict()
        self.vacancy_rate_by_city = dict()
        self.dict_lict = list()


class InputConnect:
    """Класс для ввода данных и формирования отчетности о вакансиях

    Args:
        params (tuple): Кортеж с названием файла и профессии
    """

    def __init__(self):
        """Конструктор для инициализации объекта InputConnect"""
        self.params = InputConnect.get_params()

    @staticmethod
    def get_params():
        """Статический метод для ввода данные о вакансии
        :return: Кортеж с названием файла и профессии
        """
        file_name = "vacancies_by_year.csv"  # input("Введите название файла: ")
        profession_name = "Программист"  # input("Введите название профессии: ")
        return file_name, profession_name

    @staticmethod
    def print_data_dict(self, data: DataSet):
        """Вычисляет и печатает в консоль словари со статистикой о вакансиях
        :param self: Объект класса InputConnect
        :param data: Объект класса DataSet
        """
        df = pd.read_csv(data.file_name)
        df["salary"] = df[["salary_from", "salary_to"]].mean(axis=1)
        df["published_at"] = df["published_at"].apply(lambda d: datetime(int(d[:4]), int(d[5:7]), int(d[8:10])).year)
        years = df["published_at"].unique()
        df_vacancy = df["name"].str.contains(self.params[1])

        for year in years:
            filter_by_year = df["published_at"] == year
            data.salary_by_year[year] = int(df[filter_by_year]["salary"].mean())
            data.vacancies_count_by_year[year] = len(df[filter_by_year])
            data.salary_by_profession_name[year] = int(df[df_vacancy & filter_by_year]["salary"].mean())
            data.vacancies_count_by_profession_name[year] = len(df[df_vacancy & filter_by_year])

        vacancy_rate_by_city = dict(df["area_name"].value_counts())
        for area in [k for k, v in vacancy_rate_by_city.items() if math.floor(v / len(df.index) * 100) >= 1]:
            data.salary_by_city[area] = int(
                df.loc[df['area_name'] == area, 'salary'].sum() / vacancy_rate_by_city[area])
        data.vacancy_rate_by_city = {k: round(v / len(df.index), 4) for k, v in
                                     dict(df["area_name"].value_counts()).items()}
        data.salary_by_city = dict(sorted(data.salary_by_city.items(), key=lambda item: item[1], reverse=True))

        print(f"Динамика уровня зарплат по годам: {data.salary_by_year}")
        print(f"Динамика количества вакансий по годам: {data.vacancies_count_by_year}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {data.salary_by_profession_name}")
        print(
            f"Динамика количества вакансий по годам для выбранной профессии: {data.vacancies_count_by_profession_name}")
        print(f"Уровень зарплат по городам (в порядке убывания): {dict(list(data.salary_by_city.items())[:10])}")
        print(f"Доля вакансий по городам (в порядке убывания): {dict(list(data.vacancy_rate_by_city.items())[:10])}")
        data.dict_lict = [data.salary_by_year, data.salary_by_profession_name, data.vacancies_count_by_year,
                          data.vacancies_count_by_profession_name, dict(list(data.salary_by_city.items())[:10]),
                          data.vacancy_rate_by_city]

        # data.salary_by_city = {k: v / data.vacancy_rate_by_city[k] for k, v in data.salary_by_city.items()}
        #
        # data.salary_by_city = dict(sorted(data.salary_by_city.items(), key=lambda item: item[1], reverse=True))
        # data.vacancy_rate_by_city = {k: round(v / self.count, 4) for k, v in data.vacancy_rate_by_city.items()}



class Report:
    """Класс для формирования отчетности в виде pdf, excel или png файла

    Args:
        data (list): Список словарей со статистикой о вакансиях
    """

    def __init__(self, dict_lict: list()):
        """Конструктор для инициализации объекта Report
        :param dict_lict: Список словарей со статистикой о вакансиях
        """
        self.data = dict_lict

    def generate_excel(self, profession_name):
        """Метод для генерации excel файла по названию профессии, после запуска данного метода
        файл с расширением xlsx появится в локальной директории проекта.

        :param profession_name: Название профессии
        :return: None
        """

        def as_text(value):
            """Функция, которая преобразует входное значение в тип str
            :param value: Any
            :return: str или "" Если value is None
            """
            if value is None:
                return ""
            return str(value)

        def set_max_length(worksheet):
            """Устанавливает максимальную длинну колонки в таблицу
            :param worksheet: Рабочая область таблицы
            """
            for column_cells in worksheet.columns:
                length = max(len(as_text(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        def set_format_percent(worksheet):
            """Устанавливает в 5 колонке формат отображения данных в виде процентов
            :param worksheet: Рабочая область таблицы
            """
            for i, column_cells in enumerate(worksheet.columns):
                if i == 4:
                    for cell in column_cells:
                        cell.number_format = FORMAT_PERCENTAGE_00

        def set_border_style(worksheet):
            """Устанавливает стиль границам заполненных ячеек
            :param worksheet: Рабочая область таблицы
            """
            for column_cells in worksheet.columns:
                for cell in column_cells:
                    bd = Side(style="thin", color="000000")
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        def set_headers(headers, head_range):
            """Устанавливает в первый ряд заголовки колонок
            :param headers: Список заголовок
            :param head_range: Диапазон значений для заголовок
            :return:
            """
            for i, cell in enumerate(head_range):
                cell.value = headers[i]
                cell.font = Font(size=11, b=True)

        wb = Workbook()
        sheet_1 = wb.worksheets[0]
        sheet_1.title = "Статистика по годам"
        sheet_2 = wb.create_sheet("Статистика по городам")
        headers = ["Год", "Средняя зарплата", f"Средняя зарплата - {profession_name}",
                   "Количество вакансий", f"Количество вакансий - {profession_name}"]
        set_headers(headers, sheet_1['A1':'E1'][0])

        for key in self.data[0].keys():
            sheet_1.append([key, self.data[0][key], self.data[1][key], self.data[2][key], self.data[3][key]])
        set_border_style(sheet_1)
        set_max_length(sheet_1)

        set_headers(["Город", "Уровень зарплат"], sheet_2['A1':'B1'][0])
        set_headers(["Город", "Доля вакансий"], sheet_2['D1':'E1'][0])
        sheet_2.column_dimensions['C'].width = 2
        city_keys = list(self.data[5].keys())
        for i, key in enumerate(self.data[4].keys()):
            sheet_2.append([key, self.data[4][key], None, city_keys[i], self.data[5][city_keys[i]]])

        for i, column_cells in enumerate(sheet_2.columns):
            for cell in column_cells:
                if i != 2:
                    bd = Side(style="thin", color="000000")
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        set_format_percent(sheet_2)
        set_max_length(sheet_2)
        wb.save("report.xlsx")
        return

    def generate_image(self, profession_name):
        """Метод для генерирования картинки по названию профессии с графиками
        после запуска данного метода файл с расширением .png появится в локальной директории проекта.
        :param profession_name: Название професии
        """

        def myfunc(item):
            """Фукнция, которая устанавливает символ \n в строку, если в ней имеет символ ' ' или '-'
            :param item: Строка
            """
            if item.__contains__(' '):
                return item[:item.index(' ')] + '\n' + item[item.index(' ') + 1:]
            elif item.__contains__('-'):
                return item[:item.index('-')] + '-\n' + item[item.index('-') + 1:]
            return item

        width = 0.3
        nums = np.arange(len(self.data[0].keys()))
        dx1 = nums - width / 2
        dx2 = nums + width / 2

        fig = plt.figure()
        ax = fig.add_subplot(221)
        ax.set_title("Уровень зарплат по годам")
        ax.bar(dx1, self.data[0].values(), width, label="средняя з/п")
        ax.bar(dx2, self.data[1].values(), width, label=f"з/п {profession_name.lower()}")
        ax.set_xticks(nums, self.data[0].keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis='y')

        ax = fig.add_subplot(222)
        ax.set_title("Количество вакансии по годам")
        ax.bar(dx1, self.data[2].values(), width, label="Количество вакансии")
        ax.bar(dx2, self.data[3].values(), width, label=f"Количество вакансии\n{profession_name.lower()}")
        ax.set_xticks(nums, self.data[0].keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis='y')

        ax = fig.add_subplot(223)
        ax.set_title("Уровень зарплат по городам")
        cities = list(map(myfunc, tuple(self.data[4].keys())))
        y_pos = np.arange(len(cities))
        ax.barh(y_pos, list(self.data[4].values()), align='center')
        ax.set_yticks(y_pos, labels=cities)
        ax.invert_yaxis()
        ax.grid(True, axis='x')

        ax = fig.add_subplot(224)
        ax.set_title("Доля вакансии по годам")
        labels = list(dict(list(self.data[5].items())[:10]).keys())
        labels.insert(0, "Другие")
        vals = list(dict(list(self.data[5].items())[:10]).values())
        vals.insert(0, 1 - sum(list(dict(list(self.data[5].items())[:10]).values())))
        ax.pie(vals, labels=labels, startangle=0, textprops={"fontsize": 6})
        plt.tight_layout()
        fig.set_size_inches(9.5, 7.5)
        plt.savefig("graph.png", dpi=120)
        return

    def generate_pdf(self, profession_name):
        """Метода для генерации отчетности с графиком и таблицами.
        После запуска данного метода файл с расширением .pdf появится в локальной директории проекта.

        :param profession_name: Название профессии
        """
        self.generate_excel(profession_name)
        self.generate_image(profession_name)
        name = profession_name
        image_file = "graph.png"
        book = load_workbook("report.xlsx")
        sheet_1 = book.active
        sheet_2 = book['Статистика по городам']
        for row in range(2, sheet_2.max_row + 1):
            for col in range(4, 6):
                if type(sheet_2.cell(row, col).value).__name__ == "float":
                    sheet_2.cell(row, col).value = str(round(sheet_2.cell(row, col).value * 100, 2)) + '%'

        options = {'enable-local-file-access': None}
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render({'name': name, 'image_file': image_file, 'sheet_1': sheet_1, 'sheet_2': sheet_2})
        config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)


# def main_pdf():
#     """ Функция для запуска формирования отчета
#     :return: None
#     """

dirname = r"vacancies"
files = os.listdir(dirname)

inputparam = InputConnect()
start_time = datetime.now()
dataset = DataSet(inputparam.params[0])
InputConnect.print_data_dict(inputparam, dataset)
# report = Report(dataset.dict_lict)
# report.generate_pdf(inputparam.params[1])
print(f"Total time: {datetime.now() - start_time}")

